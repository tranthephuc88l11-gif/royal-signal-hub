import streamlit as st
import google.generativeai as genai
import json
import os
import io
import re
from collections import Counter
import yt_dlp
from datetime import datetime, timedelta
from youtube_transcript_api import YouTubeTranscriptApi
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 1. CONFIG & STORAGE
# ============================================================
BASE_DATABASE = "production_database"
HISTORY_FOLDER = os.path.join(BASE_DATABASE, "history")
os.makedirs(HISTORY_FOLDER, exist_ok=True)
os.makedirs(os.path.join(BASE_DATABASE, "royal_news"), exist_ok=True)
os.makedirs(os.path.join(BASE_DATABASE, "nnt_news"), exist_ok=True)

ROYAL_CHANNELS = [
    "https://www.youtube.com/@HollywoodThrone/videos",
    "https://www.youtube.com/channel/UCaCBv2nc2AekxV5KFLCplKg/videos",
    "https://www.youtube.com/channel/UCJCCgVv5u_isRY-I5m32raA/videos",
    "https://www.youtube.com/channel/UCNK4XMtqYF95HUwC2FuekMg/videos",
    "https://www.youtube.com/channel/UCczmWNXO6oRS5IUA6iOwJow/videos",
    "https://www.youtube.com/channel/UCmxSN_yZeWuHIyUXP6gu8NA/videos",
    "https://www.youtube.com/channel/UCKArn2xo9aGL38Ocl7VSg-w/videos",
    "https://www.youtube.com/channel/UCNHQhhkNOjXAa_oHSYR75fQ/videos",
    "https://www.youtube.com/channel/UCo8FpL2u-Bmtih7wdCuhi1Q/videos",
    "https://www.youtube.com/channel/UC3UTKqCNoLTC7JyveFwnGGg/videos",
    "https://www.youtube.com/channel/UCe7KlAAv4U_npRXm62K8T-Q/videos",
    "https://www.youtube.com/channel/UCDBTjil_KwI-mXHBSg_75og/videos",
    "https://www.youtube.com/channel/UC-PPFwpQASUdYGchfYIqahw/videos",
    "https://www.youtube.com/channel/UCmcB9c2nMOjgAmaZctogKUQ/videos",
    "https://www.youtube.com/channel/UCxbtwyQ8wIn2ToPggn9l6ZA/videos",
    "https://www.youtube.com/channel/UCh37D27Avf9VH_UAripgm-g/videos",
    "https://www.youtube.com/channel/UCkEDrHRM_T3sTcrx67P_8ag/videos",
    "https://www.youtube.com/@RoyalFiles-o8k/videos",
    "https://www.youtube.com/@WindsorNotes/videos",
    "https://www.youtube.com/@crown-g6c/videos",
]

NNT_CHANNELS = [
    "https://www.youtube.com/@UnfilteredDebates2/videos",
    "https://www.youtube.com/@Flaawsometalk/videos",
    "https://www.youtube.com/@WithoutACrystalBall/videos",
]

STEPS = [
    "🔍  STEP 1 — SPY TRENDING",
    "✍️  STEP 2 — REWRITE SCRIPT",
    "📊  STEP 3 — SEO & THUMBNAIL",
    "📜  STEP 4 — HISTORY",
]

# 8-phase NNT structure for reference in prompts
NNT_STRUCTURE = """
PHASE 1 — HOOK (choose one formula):
- Shocking Statement: Bold claim → pause → "Let me explain"
- Timeline Bomb: 3-4 events with dates → "Coincidence?"
- Insider Reveal: "Someone in the room just broke their silence"
- Rhetorical Trap: Unanswerable question → list names → "Pattern?"
- Stakes Opener: Why this matters → who disappeared → "You need to watch this"

PHASE 2 — PULL QUESTION (two hundred eighty to three hundred sixty words):
Central question + three sub-questions + context (three sentences max) + The Promise

PHASE 3 — ACT ONE EVIDENCE (seven hundred twenty to eight hundred eighty words | three to four blocks):
Each block: Setup → Evidence (paraphrased) → Analysis → Social proof → Transition
Mini-Hook at minute eight (seventy to ninety words)

PHASE 4 — ACT TWO-A EVIDENCE (eight hundred ten to nine hundred ninety words | three to four blocks):
Escalate connections. Hedge language increases. Cross-reference two sources per block.
Mini-Hook at minute fourteen (seventy to ninety words)

PHASE 5 — ACT TWO-B EVIDENCE (eight hundred ten to nine hundred ninety words | three to four blocks):
Conspiracy layer. Full hedge language. "Word on the street is..." "allegedly..." "reportedly..."
Mini-Hook at minute twenty (seventy to ninety words)

PHASE 6 — ACT THREE SYNTHESIS (six hundred forty to eight hundred words | two to three blocks):
No new info. Connect all dots. Callback to central question. Pattern statement. Hedge — no hard conclusion.
Mini-Hook at minute twenty-five (seventy to ninety words)

PHASE 7 — FALSE BALANCE (four hundred thirty to five hundred thirty words):
Introduce strongest objection → Steel-man it → Quote opposing side → Rebut firmly → Reinforce narrative

PHASE 8 — CTA + CLIFFHANGER (two hundred ten to two hundred seventy words):
Partial close → Comment bait (two clear sides) → Teaser for next video
"""

# ============================================================
# 2. SESSION STATE
# ============================================================
_DEFAULTS = {
    "gemini_key": "",
    "trending_list": [],
    "selected_topic": "",
    "selected_transcript": "",
    "outline": None,
    "current_part": 0,
    "full_script_list": [],
    "seo_result": "",
    "active_channel": "👑 ROYAL NEWS",
}

for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ============================================================
# 3. HELPERS
# ============================================================
def call_ai(prompt: str) -> str:
    key = st.session_state.gemini_key
    if not key:
        return "❌ MISSING API KEY"
    genai.configure(api_key=key)
    try:
        model = genai.GenerativeModel("gemini-3-flash-preview")
        return model.generate_content(prompt).text
    except Exception as e:
        return f"❌ Gemini Error: {e}"


def get_yt_trending(url: str, days: int = 7, min_views: int = 5000):
    cutoff = datetime.now() - timedelta(days=days)
    ydl_opts = {
        "quiet": True,
        "extract_flat": False,
        "playlistend": 50,
        "ignoreerrors": True,
        "no_warnings": True,
        "nocheckcertificate": True,
        "http_headers": {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        },
    }
    results = []
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(url, download=False)
            if info and "entries" in info:
                for e in info["entries"]:
                    if not e:
                        continue
                    # Filter ngày thủ công
                    upload_date = e.get("upload_date", "")
                    if upload_date and len(upload_date) == 8:
                        try:
                            upload_dt = datetime.strptime(upload_date, "%Y%m%d")
                            if upload_dt < cutoff:
                                continue
                        except:
                            pass
                    # Filter views
                    if e.get("view_count", 0) < min_views:
                        continue
                    results.append({
                        "title": e.get("title", "Unknown"),
                        "id": e.get("id", ""),
                        "views": e.get("view_count", 0),
                        "comments": e.get("comment_count", 0) or 0,
                        "url": f"https://www.youtube.com/watch?v={e.get('id','')}",
                        "channel": e.get("uploader", "Unknown"),
                        "upload_date": upload_date,
                    })
    except:
        pass
    return results


STOPWORDS = set("""
a an the and or but for nor so yet of in on at to from by with as is are was were
be been being this that these those it its his her their our your my i you he she
we they will would can could should shall may might must do does did not no
""".split())


def extract_trending_keywords(videos, top_n=20):
    word_counter = Counter()
    phrase_counter = Counter()
    for v in videos:
        title = v["title"].lower()
        title_clean = re.sub(r"[^a-z0-9\s]", " ", title)
        words = [w for w in title_clean.split() if len(w) > 2 and w not in STOPWORDS]
        word_counter.update(words)
        for i in range(len(words) - 1):
            phrase_counter[f"{words[i]} {words[i+1]}"] += 1
    top_words = word_counter.most_common(top_n)
    top_phrases = [p for p in phrase_counter.most_common(top_n) if p[1] > 1]
    return top_words, top_phrases


def export_videos_to_excel(videos, channel_name):
    wb = Workbook()

    # Sheet 1: Videos
    ws = wb.active
    ws.title = "Videos"
    headers = ["Title", "Channel", "Views", "Comments", "Upload Date", "Video URL"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F2937")
        c.alignment = Alignment(horizontal="center")

    for v in videos:
        upload_fmt = v.get("upload_date", "")
        if upload_fmt and len(upload_fmt) == 8:
            upload_fmt = f"{upload_fmt[:4]}-{upload_fmt[4:6]}-{upload_fmt[6:]}"
        ws.append([
            v["title"], v.get("channel", ""), v["views"],
            v.get("comments", 0), upload_fmt, v["url"],
        ])

    widths = [55, 22, 12, 12, 14, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Sheet 2: Trending Keywords
    ws2 = wb.create_sheet("Trending Keywords")
    top_words, top_phrases = extract_trending_keywords(videos)

    ws2["A1"] = "Top Keywords (single words)"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2.append(["Keyword", "Frequency"])
    for c in ws2[2]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F2937")
    for word, count in top_words:
        ws2.append([word, count])

    start_row = len(top_words) + 4
    ws2.cell(row=start_row, column=1, value="Top Keyword Phrases (2-word combos)").font = Font(bold=True, size=12)
    ws2.cell(row=start_row + 1, column=1, value="Phrase")
    ws2.cell(row=start_row + 1, column=2, value="Frequency")
    for c in ws2[start_row + 1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="1F2937")
    for i, (phrase, count) in enumerate(top_phrases):
        ws2.cell(row=start_row + 2 + i, column=1, value=phrase)
        ws2.cell(row=start_row + 2 + i, column=2, value=count)

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def get_transcript(v_id: str) -> str:
    try:
        ts_list = YouTubeTranscriptApi.list_transcripts(v_id)
        try:
            ts = ts_list.find_transcript(["en"])
        except:
            ts = ts_list.find_generated_transcript(["en"])
        return " ".join([t["text"] for t in ts.fetch()])
    except:
        return ""


def save_to_history(topic: str, content: str, channel_tag: str):
    fname = f"{HISTORY_FOLDER}/{datetime.now().strftime('%Y%m%d_%H%M%S')}_{channel_tag}.json"
    with open(fname, "w", encoding="utf-8") as f:
        json.dump({
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "topic": topic,
            "content": content,
            "channel": channel_tag,
        }, f, ensure_ascii=False, indent=4)


def reset_project():
    for k in ["trending_list", "selected_topic", "selected_transcript",
              "outline", "current_part", "full_script_list", "seo_result"]:
        st.session_state[k] = _DEFAULTS[k]
    st.rerun()


# ============================================================
# 4. PAGE CONFIG & CSS
# ============================================================
st.set_page_config(page_title="AI Production Hub", page_icon="👑", layout="wide")

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #0f1117; }
[data-testid="stSidebar"] { background: #16181f; border-right: 1px solid #2a2d3a; }

.hub-title { font-size:1.2rem; font-weight:700; color:#f0c040; margin-bottom:.2rem; }
.hub-version { font-size:.68rem; color:#555; margin-bottom:1rem; }
.sec-label {
    font-size:.6rem; font-weight:700; letter-spacing:.12em;
    color:#666; text-transform:uppercase; margin-bottom:.3rem; margin-top:.8rem;
}

/* Channel tabs */
.chan-tab {
    display:flex; align-items:center; gap:.5rem;
    padding:.55rem .8rem; border-radius:8px; cursor:pointer;
    border:1px solid transparent; margin-bottom:.3rem;
    font-size:.88rem; font-weight:600; transition:all .15s;
}
.chan-tab.royal { color:#f0c040; border-color:#3a3010; background:#1e1a08; }
.chan-tab.nnt   { color:#a78bfa; border-color:#2d1f5e; background:#130d2e; }
.chan-tab.royal.active { border-color:#f0c040; background:#2a2208; }
.chan-tab.nnt.active   { border-color:#a78bfa; background:#1e1545; }

/* Steps */
div[data-testid="stRadio"] label {
    font-size:.83rem !important; color:#bbb !important;
    padding:.32rem .5rem !important; border-radius:6px !important;
    display:block !important;
}
div[data-testid="stRadio"] label:hover { background:#1e2130 !important; color:#eee !important; }

/* Section header */
.sec-header {
    border-left:3px solid #f0c040; padding-left:.75rem; margin-bottom:1.2rem;
}
.sec-header.nnt { border-left-color:#a78bfa; }
.sec-header h2 { font-size:1.25rem; font-weight:700; color:#f0f0f0; margin:0; }
.sec-header p  { font-size:.78rem; color:#666; margin:.1rem 0 0; }

/* Tag badge */
.tag-royal {
    display:inline-block; background:#2a2208; color:#f0c040;
    border:1px solid #f0c040; font-size:.65rem; font-weight:700;
    padding:.15rem .5rem; border-radius:99px; margin-right:.4rem;
    letter-spacing:.06em;
}
.tag-nnt {
    display:inline-block; background:#1e1545; color:#a78bfa;
    border:1px solid #a78bfa; font-size:.65rem; font-weight:700;
    padding:.15rem .5rem; border-radius:99px; margin-right:.4rem;
    letter-spacing:.06em;
}

/* Video card */
.vcard {
    background:#16181f; border:1px solid #2a2d3a; border-radius:10px;
    padding:.8rem 1rem; margin-bottom:.5rem;
}
.vcard-title { font-size:.88rem; font-weight:600; color:#eee; margin-bottom:.2rem; }
.vcard-meta  { font-size:.72rem; color:#666; }

/* Part chip */
.part-chip {
    display:inline-block; background:#1e3a5f; color:#60aaff;
    font-size:.68rem; font-weight:700; letter-spacing:.08em;
    padding:.18rem .55rem; border-radius:99px; margin-bottom:.4rem;
    text-transform:uppercase;
}
.part-chip.nnt { background:#2d1f5e; color:#a78bfa; }

/* Progress */
.prog-bg { background:#1e2130; border-radius:99px; height:7px; overflow:hidden; margin-bottom:1rem; }
.prog-fill { background:linear-gradient(90deg,#f0c040,#e07b20); height:100%; border-radius:99px; }
.prog-fill.nnt { background:linear-gradient(90deg,#a78bfa,#7c3aed); }

hr { border-color:#2a2d3a !important; margin:.8rem 0 !important; }
textarea, input[type="text"], input[type="password"] {
    background:#12151e !important; border:1px solid #2a2d3a !important;
    color:#eee !important; border-radius:8px !important;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# 5. SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown('<div class="hub-title">👑 PRODUCTION HUB</div>', unsafe_allow_html=True)
    st.markdown('<div class="hub-version">AI Script Factory · v26</div>', unsafe_allow_html=True)

    # API Key
    key_input = st.text_input(
        "Gemini API Key", type="password", placeholder="AIza...",
        value=st.session_state.gemini_key
    )
    if key_input != st.session_state.gemini_key:
        st.session_state.gemini_key = key_input
    if st.session_state.gemini_key:
        st.success("API key loaded ✅")
    else:
        st.warning("Enter API key 🔑")

    st.divider()

    # Channel selector
    st.markdown('<div class="sec-label">Channel</div>', unsafe_allow_html=True)

    col_r, col_n = st.columns(2)
    with col_r:
        royal_active = "active" if st.session_state.active_channel == "👑 ROYAL NEWS" else ""
        if st.button("👑 ROYAL NEWS", use_container_width=True,
                     type="primary" if royal_active else "secondary"):
            if st.session_state.active_channel != "👑 ROYAL NEWS":
                st.session_state.active_channel = "👑 ROYAL NEWS"
                reset_project()
    with col_n:
        nnt_active = "active" if st.session_state.active_channel == "🎭 NNT NEWS" else ""
        if st.button("🎭 NNT NEWS", use_container_width=True,
                     type="primary" if nnt_active else "secondary"):
            if st.session_state.active_channel != "🎭 NNT NEWS":
                st.session_state.active_channel = "🎭 NNT NEWS"
                reset_project()

    st.markdown(f"**Active:** `{st.session_state.active_channel}`")

    st.divider()

    # Steps
    st.markdown('<div class="sec-label">Workflow</div>', unsafe_allow_html=True)
    menu = st.radio("Steps", STEPS, key="menu_radio", label_visibility="collapsed")

    st.divider()

    # Progress bar
    curr_part = st.session_state.current_part
    if st.session_state.outline and 1 <= curr_part <= 6:
        pct = int((curr_part - 1) / 6 * 100)
        color_class = "nnt" if "NNT" in st.session_state.active_channel else ""
        st.markdown(f"""
        <div style="font-size:.72rem;color:#888;margin-bottom:.3rem;">
            Script Progress — Part {curr_part-1} of 6
        </div>
        <div class="prog-bg">
            <div class="prog-fill {color_class}" style="width:{pct}%"></div>
        </div>
        """, unsafe_allow_html=True)

    if st.button("↺  Reset Project", type="secondary", use_container_width=True):
        reset_project()

# ============================================================
# 6. MAIN CONTENT
# ============================================================
api_ready = bool(st.session_state.gemini_key)
is_nnt = "NNT" in st.session_state.active_channel
channel_tag = "NNT" if is_nnt else "ROYAL"
accent = "#a78bfa" if is_nnt else "#f0c040"
header_class = "sec-header nnt" if is_nnt else "sec-header"
chip_class = "part-chip nnt" if is_nnt else "part-chip"


def section_header(title: str, subtitle: str):
    st.markdown(f"""
    <div class="{header_class}">
        <h2>{title}</h2>
        <p>{subtitle}</p>
    </div>
    """, unsafe_allow_html=True)


# ── STEP 1: SPY TRENDING ─────────────────────────────────
if menu == STEPS[0]:
    section_header("Viral Spy", f"{st.session_state.active_channel} · Last 7 days · ≥ 10,000 views")

    channels = NNT_CHANNELS if is_nnt else ROYAL_CHANNELS

    # Settings row
    cfg1, cfg2, cfg3 = st.columns([1.5, 1.5, 1])
    with cfg1:
        scan_mode = st.radio("Mode", ["All channels", "Single channel"],
                             horizontal=True, label_visibility="collapsed")
    with cfg2:
        min_views = st.select_slider(
            "Min views", options=[10000, 20000, 50000, 100000],
            value=10000, label_visibility="collapsed",
            format_func=lambda x: f"≥ {x:,} views"
        )
    with cfg3:
        days_back = st.select_slider(
            "Days", options=[3, 7, 14, 30],
            value=7, label_visibility="collapsed",
            format_func=lambda x: f"Last {x}d"
        )

    col_sel, col_btn = st.columns([3, 1])
    with col_sel:
        if scan_mode == "Single channel":
            target = st.selectbox("Channel", channels, label_visibility="collapsed")
        else:
            st.caption(f"Scanning all **{len(channels)}** channels · last {days_back} days · ≥ {min_views:,} views")
    with col_btn:
        if st.button("▶  Scan", type="primary", use_container_width=True):
            if scan_mode == "Single channel":
                with st.spinner("Scanning..."):
                    st.session_state.trending_list = get_yt_trending(target, days=days_back, min_views=min_views)
            else:
                all_results = []
                progress = st.progress(0)
                status = st.empty()
                for i, ch in enumerate(channels):
                    status.text(f"Scanning {i+1}/{len(channels)}: {ch.split('/')[-2]}")
                    all_results.extend(get_yt_trending(ch, days=days_back, min_views=min_views))
                    progress.progress((i + 1) / len(channels))
                status.empty()
                progress.empty()
                # Deduplicate by video id
                seen = set()
                unique = []
                for v in all_results:
                    if v["id"] not in seen:
                        seen.add(v["id"])
                        unique.append(v)
                unique.sort(key=lambda v: v["views"], reverse=True)
                st.session_state.trending_list = unique

    if st.session_state.trending_list:
        videos = st.session_state.trending_list
        col_count, col_export = st.columns([4, 1])
        with col_count:
            st.caption(f"Found **{len(videos)}** videos")
        with col_export:
            excel_buf = export_videos_to_excel(videos, st.session_state.active_channel)
            st.download_button(
                "📊  Export to Excel",
                data=excel_buf,
                file_name=f"trending_{channel_tag}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        st.divider()
        for v in videos:
            col_info, col_act = st.columns([5, 1])
            with col_info:
                st.markdown(f"""
                <div class="vcard">
                    <div class="vcard-title">🔥 {v['title']}</div>
                    <div class="vcard-meta">👁 {v['views']:,} views &nbsp;·&nbsp;
                    💬 {v.get('comments', 0):,} comments &nbsp;·&nbsp;
                    <a href="{v['url']}" target="_blank" style="color:{accent}">Open ↗</a></div>
                </div>
                """, unsafe_allow_html=True)
            with col_act:
                if st.button("Select", key=f"sel_{v['id']}", use_container_width=True):
                    with st.spinner("Fetching transcript..."):
                        st.session_state.selected_topic = v["title"]
                        st.session_state.selected_transcript = get_transcript(v["id"])
                    st.success("Loaded → go to Step 2")


# ── STEP 2: REWRITE SCRIPT ───────────────────────────────
elif menu == STEPS[1]:
    if is_nnt:
        section_header("Script Master", "NNT NEWS · 8-Phase celebrity drama structure · Gemini 3")
    else:
        section_header("Rewrite Master", "ROYAL NEWS · 6-part structure · 5,000 words · Gemini 3")

    c1, c2 = st.columns([1, 1])
    with c1:
        st.session_state.selected_topic = st.text_input(
            "Project name / topic", value=st.session_state.selected_topic
        )
    with c2:
        st.session_state.selected_transcript = st.text_area(
            "Transcript source", value=st.session_state.selected_transcript, height=120
        )

    st.divider()

    if not api_ready:
        st.warning("Enter your Gemini API key in the sidebar first.", icon="🔑")

    # ── ROYAL: 6-part outline flow ──
    if not is_nnt:
        if st.session_state.outline is None:
            st.info("Generate outline first, then write each part.", icon="ℹ️")
            can_go = api_ready and bool(st.session_state.selected_transcript)
            if st.button("Generate Outline →", type="primary", disabled=not can_go):
                prompt_o = f"""
ACT AS A MASTER CONTENT CREATOR specializing in royal drama and high-retention YouTube scripts.
Task: Create a detailed OUTLINE to rewrite this transcript into a five-thousand-word viral YouTube script.

STRICT RULES:
1. Divide into EXACTLY SIX PARTS.
2. Assign a specific word count per part (total = five thousand words).
3. NO DIGITS — write all numbers as words (e.g. 'eight hundred' not 800).
4. ENGLISH ONLY.

Topic: {st.session_state.selected_topic}
Transcript: {st.session_state.selected_transcript[:3500]}
"""
                with st.spinner("Generating outline..."):
                    res = call_ai(prompt_o)
                    if "❌" not in res:
                        st.session_state.outline = res
                        st.session_state.current_part = 1
                        st.rerun()
                    else:
                        st.error(res)
        else:
            curr = st.session_state.current_part
            with st.expander("📋 View Outline", expanded=(curr == 1)):
                st.markdown(st.session_state.outline)
            st.divider()

            if curr <= 6:
                col_s, col_a = st.columns([3, 1])
                with col_s:
                    st.markdown(f'<div class="{chip_class}">Part {curr} / 6</div>', unsafe_allow_html=True)
                    st.progress(int((curr - 1) / 6 * 100) / 100)
                with col_a:
                    write_btn = st.button(f"Write Part {curr} →", type="primary",
                                          use_container_width=True, disabled=not api_ready)
                if write_btn:
                    prompt_w = f"""
You are a master YouTube scriptwriter specializing in royal drama and high-retention content.

FULL OUTLINE:
{st.session_state.outline}

TASK: Write PART {curr} of six for the script about: "{st.session_state.selected_topic}"

REQUIREMENTS:
1. Write ONLY Part {curr} — do not write other parts.
2. Follow the word count assigned to Part {curr} in the outline (around eight hundred words).
3. NO DIGITS — write all numbers as words (e.g. 'twenty-twenty-five' not two thousand and twenty-five).
4. ENGLISH ONLY.
5. Style: cinematic, dramatic, high-stakes royal storytelling.
6. PURE NARRATION ONLY — no image suggestions, no B-roll notes, no [visual], no [cut to], no brackets of any kind.
"""
                    with st.spinner(f"Writing Part {curr}..."):
                        result = call_ai(prompt_w)
                        if "❌" not in result:
                            st.session_state.full_script_list.append(f"## PART {curr}\n\n{result}")
                            if curr == 6:
                                save_to_history(
                                    st.session_state.selected_topic,
                                    "\n\n".join(st.session_state.full_script_list),
                                    "ROYAL"
                                )
                            st.session_state.current_part += 1
                            st.rerun()
                        else:
                            st.error(result)

            if st.session_state.full_script_list:
                st.divider()
                st.caption(f"{len(st.session_state.full_script_list)} part(s) written")
                for p in st.session_state.full_script_list:
                    st.markdown(p)
                    st.divider()

            if curr > 6:
                st.success("✅ Script complete — saved to History.")
                st.download_button("⬇  Download Script (.md)",
                                   "\n\n".join(st.session_state.full_script_list),
                                   file_name=f"{st.session_state.selected_topic or 'script'}.md",
                                   mime="text/markdown", type="primary")

    # ── NNT: 8-phase flow ──
    else:
        if st.session_state.outline is None:
            st.info("Generate 8-phase outline first, then write each phase.", icon="ℹ️")
            can_go = api_ready and bool(st.session_state.selected_transcript)
            if st.button("Generate 8-Phase Outline →", type="primary", disabled=not can_go):
                prompt_o = f"""
You are a master YouTube scriptwriter specializing in celebrity drama, conspiracy, and high-retention content.

Task: Create a detailed OUTLINE based on the eight-phase structure below to rewrite this transcript into a full viral YouTube script.

EIGHT-PHASE STRUCTURE TO FOLLOW:
{NNT_STRUCTURE}

STRICT RULES:
1. Create an outline for ALL EIGHT PHASES.
2. For each phase, specify: key points to cover, hook formula to use (Phase one), word count target.
3. NO DIGITS — write all numbers as words.
4. ENGLISH ONLY.

Topic: {st.session_state.selected_topic}
Transcript: {st.session_state.selected_transcript[:3500]}
"""
                with st.spinner("Generating 8-phase outline..."):
                    res = call_ai(prompt_o)
                    if "❌" not in res:
                        st.session_state.outline = res
                        st.session_state.current_part = 1
                        st.rerun()
                    else:
                        st.error(res)
        else:
            curr = st.session_state.current_part
            with st.expander("📋 View 8-Phase Outline", expanded=(curr == 1)):
                st.markdown(st.session_state.outline)
            st.divider()

            phase_labels = [
                "Phase 1 — Hook",
                "Phase 2 — Pull Question",
                "Phase 3 — Act One Evidence",
                "Phase 4 — Act Two-A Evidence",
                "Phase 5 — Act Two-B Evidence",
                "Phase 6 — Act Three Synthesis",
                "Phase 7 — False Balance",
                "Phase 8 — CTA + Cliffhanger",
            ]

            if curr <= 8:
                col_s, col_a = st.columns([3, 1])
                with col_s:
                    label = phase_labels[curr - 1] if curr <= 8 else ""
                    st.markdown(f'<div class="{chip_class}">{label}</div>', unsafe_allow_html=True)
                    st.progress(int((curr - 1) / 8 * 100) / 100)
                with col_a:
                    write_btn = st.button(f"Write Phase {curr} →", type="primary",
                                          use_container_width=True, disabled=not api_ready)

                if write_btn:
                    prompt_w = f"""
You are a master YouTube scriptwriter specializing in celebrity drama and high-retention content.

FULL OUTLINE:
{st.session_state.outline}

EIGHT-PHASE STRUCTURE REFERENCE:
{NNT_STRUCTURE}

TASK: Write PHASE {curr} ({phase_labels[curr-1]}) of the script about: "{st.session_state.selected_topic}"

REQUIREMENTS:
1. Write ONLY Phase {curr} — follow its structure and word count from the outline exactly.
2. NO DIGITS — write all numbers as words (e.g. 'twenty-twenty-five').
3. ENGLISH ONLY.
4. Style: dramatic, conspiratorial, high-stakes celebrity storytelling. Hedge language where needed.
5. PURE NARRATION ONLY — no image suggestions, no [visual], no [cut to], no brackets of any kind.
6. Use the hook formula specified in the outline for Phase one.
"""
                    with st.spinner(f"Writing {phase_labels[curr-1]}..."):
                        result = call_ai(prompt_w)
                        if "❌" not in result:
                            st.session_state.full_script_list.append(
                                f"## {phase_labels[curr-1].upper()}\n\n{result}"
                            )
                            if curr == 8:
                                save_to_history(
                                    st.session_state.selected_topic,
                                    "\n\n".join(st.session_state.full_script_list),
                                    "NNT"
                                )
                            st.session_state.current_part += 1
                            st.rerun()
                        else:
                            st.error(result)

            if st.session_state.full_script_list:
                st.divider()
                st.caption(f"{len(st.session_state.full_script_list)} phase(s) written")
                for p in st.session_state.full_script_list:
                    st.markdown(p)
                    st.divider()

            max_parts = 8 if is_nnt else 6
            if curr > max_parts:
                st.success("✅ Full script complete — saved to History.")
                st.download_button("⬇  Download Script (.md)",
                                   "\n\n".join(st.session_state.full_script_list),
                                   file_name=f"{st.session_state.selected_topic or 'script'}.md",
                                   mime="text/markdown", type="primary")


# ── STEP 3: SEO & THUMBNAIL ──────────────────────────────
elif menu == STEPS[2]:
    if is_nnt:
        section_header("SEO & Thumbnail", "NNT NEWS · AI-generated titles · description · hashtags")
    else:
        section_header("SEO & Thumbnail", "ROYAL NEWS · 5 titles · thumbnail brief · full description")

    if not api_ready:
        st.warning("Enter your Gemini API key in the sidebar first.", icon="🔑")

    col_a, col_b = st.columns(2)
    with col_a:
        titles = st.text_area("Paste competitor titles here",
                              height=200, placeholder="One title per line...")
    with col_b:
        script_box = st.text_area("Your script (auto-filled from Step 2)",
                                  value="\n\n".join(st.session_state.full_script_list), height=200)

    st.divider()

    if st.button("🚀  Generate SEO Package", type="primary", disabled=not api_ready):
        if is_nnt:
            prompt_seo = f"""
You are an expert YouTube SEO strategist for celebrity drama and gossip content.

Analyze these competitor titles:
{titles}

Based on the script below, generate a complete SEO package:

TASK ONE — TITLES:
Create five viral titles in the same emotional, conspiratorial style as the competitor titles above.
Follow this formula: [Celebrity Name] + [Emotional Verb] + "As" + [Specific Shocking Event]
Examples of style: shocking, dramatic, present-tense, curiosity-gap driven.

TASK TWO — THUMBNAIL BRIEF:
Write a thumbnail composition brief including:
- Main visual element
- Text overlay (max six words, all caps)
- Color mood and contrast
- Emotional expression of subject

TASK THREE — VIDEO DESCRIPTION:
Write a full YouTube description using this structure:
[Two to three dramatic sentences summarizing the video — based on: {script_box[:400]}]
• [Key revelation one]
• [Key revelation two]
• [Key revelation three]
👉 Subscribe for more celebrity deep dives and untold stories.

📍 Disclaimer
This video is for entertainment and commentary purposes only. All claims are based on publicly available information and alleged reports. We do not verify all allegations. Content is examined under YouTube Fair Use.

TASK FOUR — HASHTAGS:
Generate fifteen relevant hashtags for celebrity drama content on this topic.

ENGLISH ONLY. NO DIGITS — write all numbers as words.

Script reference: {script_box[:600]}
"""
        else:
            prompt_seo = f"""
You are an expert YouTube SEO strategist for royal family content.

Analyze these competitor titles:
{titles}

TASK ONE — TITLES: Create five new viral titles in the same royal drama style.
TASK TWO — THUMBNAIL: Write a thumbnail composition brief (main visual, text overlay, color mood).
TASK THREE — DESCRIPTION using this exact template:

[Two to three dramatic sentences based on: {script_box[:400]}]
• [Key point one]
• [Key point two]
• [Key point three]
👉 Don't miss our deep dives into royal truth and tradition.
Subscribe: https://www.youtube.com/@RoyalSignal-1 and turn on the bell!

📍 Disclaimer
Independent commentary and analysis for discussion. We do not verify allegations. Examined under YouTube Fair Use.
#KateMiddleton #QueenCamilla #RoyalNews #BreakingNews #PrincessOfWales #RoyalFamily #HouseOfWindsor #RoyalExpert #KingCharlesIII #BritishMonarchy

ENGLISH ONLY.
"""
        with st.spinner("Generating SEO package..."):
            st.session_state.seo_result = call_ai(prompt_seo)

    if st.session_state.seo_result:
        st.divider()
        st.markdown(st.session_state.seo_result)
        st.download_button("⬇  Download SEO Package", st.session_state.seo_result,
                           file_name="seo_package.txt", mime="text/plain")


# ── STEP 4: HISTORY ──────────────────────────────────────
elif menu == STEPS[3]:
    section_header("Script History", "All channels · Shared library")

    files = sorted(
        [f for f in os.listdir(HISTORY_FOLDER) if f.endswith(".json")], reverse=True
    )

    if not files:
        st.info("No scripts saved yet.", icon="📂")
    else:
        # Filter
        filter_col, _ = st.columns([2, 4])
        with filter_col:
            filter_opt = st.selectbox("Filter by channel", ["All", "ROYAL", "NNT"])

        st.caption(f"{len(files)} total script(s)")
        st.divider()

        for fn in files:
            fpath = os.path.join(HISTORY_FOLDER, fn)
            with open(fpath, "r", encoding="utf-8") as fh:
                data = json.load(fh)

            tag = data.get("channel", "ROYAL")
            if filter_opt != "All" and tag != filter_opt:
                continue

            tag_html = f'<span class="tag-{tag.lower()}">{tag}</span>'
            col_exp, col_dl, col_del = st.columns([6, 1, 1])
            with col_exp:
                with st.expander(f"📅 {data['date']}  ·  🎬 {data['topic']}"):
                    st.markdown(tag_html, unsafe_allow_html=True)
                    st.markdown(data["content"])
            with col_dl:
                st.download_button("⬇", data["content"],
                                   file_name=f"{data['topic']}.md",
                                   mime="text/markdown",
                                   key=f"dl_{fn}", use_container_width=True)
            with col_del:
                if st.button("🗑", key=f"del_{fn}", use_container_width=True):
                    os.remove(fpath)
                    st.toast("Deleted.")
                    st.rerun()
